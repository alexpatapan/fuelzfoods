
import pandas as pd
import datetime
import numpy as np
import openpyxl as oxl
from random import seed
import time
import random
import sys
from datetime import date, timedelta


def getCustoms():
    return Customs

def getPOrds():
    return POrds

def getCustomsLen():
    return numCustom

def getCurrentWeek():
    if (datetime.date.isocalendar(datetime.date.today())[1] < 10):
        week = '0' + str(datetime.date.isocalendar(datetime.date.today())[1])
    else:
        week = str(datetime.date.isocalendar(datetime.date.today())[1]) 
    return str(int(datetime.date.today().year) - 2000) + week

def getLastWeek():
    #Current Week
    currentYear = int(datetime.date.today().year) - 2000
    currentWeek= datetime.date.isocalendar(datetime.date.today())[1]

    if (currentWeek == 1):
        return str(currentYear-1) + str(weeks_for_year(currentYear-1)) 
        print(PAST_ORDER_FILE)
    elif (currentWeek < 11) :
        return str(currentYear) + '0' + str(currentWeek - 1)
    else:
        return str(currentYear) + str(currentWeek - 1) 

def weeks_for_year(year):
    last_week = date(year, 12, 28)
    return last_week.isocalendar()[1]

def main(ORDER_FILE):
    random.seed();
    np.set_printoptions(threshold=sys.maxsize)
    pd.set_option('display.max_colwidth', 30)

    #FILES
    PAST_ORDER_FILE = getLastWeek() + '.xlsx'
    MENU_FILE = 'Menu.xlsx'
    INGREDIENTS_FILE = 'Ingredient and meal costing1.xlsx'

    # ----------------------------------------------------------------------
    # Creating data frame from csv of todays date 

    CSVOrderFile=ORDER_FILE
    df=pd.read_csv(CSVOrderFile)


    # ----------------------------------------------------------------------
    # Searching orders for meal plans
    sub ='Plan'
    df['Plan?']=df["Item Name"].str.find(sub)  
    c=0      
    for entry in df['Plan?']:
        if entry <= 0:
            df.at[c,'Plan?']=0
        else:
            df.at[c,'Plan?']=1
        c=c+1

    # Text searching for meals per day from plans (MPD)      
    Meals_per_day = df["Product Variation"].str.extract(pat = '(: .)')
    for i in range(len(Meals_per_day)):
        if str(df.at[i,'Product Variation'])[0]!='M':
            Meals_per_day.iat[i,0] = 0
        else:
            Meals_per_day.iat[i,0] = int(str(Meals_per_day.iat[i,0])[2]) 
    df['MPD']=Meals_per_day


    # ----------------------------------------------------------------------
    # Creating classic/asian Order table from MPD

    """
    Method to search table for plan type and add the number of meals requested
    to the appropriate column
        POrds - table
        planType - The subscription plan (E.g. Asian Plan / Subscription)
        columnName - The POrds column to add the meals to
    """
    def createPOrds(POrds, planType, columnName, i) :
        if df.at[i,'Item Name'] == planType:
            for s in range(0, df.at[i,'MPD'] * 5):
                POrds.at[s,columnName] += 1
            
    global POrds
    POrds=pd.DataFrame(np.zeros((10, 4)), columns=['Classic', 'numC', 'Asian', 'numA'], dtype=object)

    for i in range(0, len(df['Order Number'])):
        
        # Repeat for number of quantities
        for k in range(0, df.at[i, 'Quantity']):
            # Normal meal plans
            createPOrds(POrds, "Subscription Plan", "numC", i)
            createPOrds(POrds, "Asian Plan", "numA", i)

            # Fusion Meals
            if df.at[i,'Item Name'] == 'Fusion Meal Plan' and df.at[i,'MPD'] == 1:
                if (int(getCurrentWeek()) % 2) == 0:
                    for s in range(0,2):
                        POrds.at[s,'numA'] += 1
                    for s in range(0,3):
                        POrds.at[s,'numC'] += 1
                else:
                    for s in range(0,3):
                        POrds.at[s,'numA'] += 1
                    for s in range(0,2):
                        POrds.at[s,'numC'] += 1          

            if df.at[i,'Item Name']=='Fusion Meal Plan' and df.at[i,'MPD']==2:
                for s in range(0, 5):
                    POrds.at[s,'numA'] += 1
                    POrds.at[s,'numC'] += 1

    # ----------------------------------------------------------------------
    # Loading previous weeks orders into dataframes (HistoricalWeek1 and HistoricalWeek2)

    HistoricalWeek1=pd.DataFrame(np.zeros((10, 4)), columns=['Classic', 'numC', 'Asian', 'numA'], dtype=object)


    OMR=oxl.load_workbook(filename=PAST_ORDER_FILE)
    sheet = OMR['Sheet1']

    def populateHistoricalOrders(HistoricalWeek, excelColumn1, excelColumn2, orderColumn1, orderColumn2):
        inc = 0
        for i in range(2, 12):
            HistoricalWeek.at[inc, orderColumn1]=sheet[excelColumn1 + str(i)].value
            HistoricalWeek.at[inc, orderColumn2]=sheet[excelColumn2 + str(i)].value
            inc += 1
    weeknum = 0
    populateHistoricalOrders(HistoricalWeek1, 'A', 'B', 'Classic', 'numC')
    populateHistoricalOrders(HistoricalWeek1, 'C', 'D', 'Asian', 'numA')

    # ----------------------------------------------------------------------
    # Make list of custom order dishes for this week (cusomOrders)
    
    customOrders = []
    customOrdersQuantity = []
    for i, j in enumerate(df['Plan?']):
        if j == 0:
            customOrders.append(df.at[i, 'Item Name'])
            customOrdersQuantity.append(df.at[i, 'Quantity'])

    #----------------------------------------------------------------------
    # Loading whole menu into a dictionary (mealMenu), Types: V=Veg, R=Red meant, C=Chicken
    OMR=oxl.load_workbook(filename=MENU_FILE)
    sheet=OMR['Menus']

    data=sheet.values
    data2 = list(data)[1:]
    MenuImport = pd.DataFrame(data2, columns=["ID", 'Menu','Dish','Type'])

    mealMenu = {}
    for i in (range(43)):
        mealMenu[MenuImport.at[i, 'Dish']] = MenuImport.at[i, 'Menu']

    # ----------------------------------------------------------------------
    #Find number of meals needed to be made (numClassics, numAsian)
    numClassics = -1
    numAsian = -1
    for i in (range(10)):
        if (POrds.at[i, 'numA'] == 0 and numClassics == -1):
            numClassics = i
        if (POrds.at[i, 'numC'] == 0 and numAsian == -1):
            numAsian = i

    # ----------------------------------------------------------------------
    #Calculate most optimal dish!
    completedClassics = 0 #num optimal dishes we have found
    completedAsian = 0

    #Possibly put custom order dishes which haven't been made last week into the meal packs (50% chance)
    for meal in customOrders:
        if (random.random() < 0.5):
            continue #Skip this meal
        if (mealMenu[meal] == 'A'):
            if not (meal in iter(HistoricalWeek1['Asian']) and not (meal in iter(POrds['Asian']))):
                POrds.at[completedAsian, 'Asian'] = meal
                completedAsian += 1
        else:
            if not (meal in iter(HistoricalWeek1['Classic']) and not (meal in iter(POrds['Classic']))):
                POrds.at[completedClassics, 'Classic'] = meal
                completedClassics += 1

    #At this point we have (possibly) added custom orders -> Pick random meal which hasn't been made in last week
    """
    Check if meal was made last week

    num - to check if in first 5 or 10
    """
    def inLastWeek(meal, menuType, num):
        for i in (range(num)):
            if (meal == HistoricalWeek1.at[i, menuType]):
                return True
        return False

    """
    Check if meal is already in POrds
    """
    def inPOrds(meal, menuType):
        for i in (range(10)):
            if (meal == POrds.at[i, menuType]):
                return True
        return False
                
    def pickRandomMeal(menuName, menuType, numMeals, completedMeals):
        while (completedMeals <= 4):
            mealNum = random.randint(1, 42)
            
            #check if meal is of correct menu type
            if (MenuImport.at[mealNum, 'Menu'] != menuType):
                continue
            meal = MenuImport.at[ mealNum, 'Dish']
            #check if meal wasnt in first 5 of last week
            if (not inLastWeek(meal, menuName, 5) and (not inPOrds(meal, menuName)) and (not meal in customOrders)):
                POrds.at[completedMeals, menuName] = meal
                completedMeals += 1

        while (completedMeals > 4 and completedMeals <= 9):
            numMealsInLastWeek = 0
            mealNum = random.randint(1,42)
            
            #check if meal is of correct menu type
            if (MenuImport.at[mealNum, 'Menu'] != menuType):
                continue
            meal = MenuImport.at[mealNum, 'Dish']
            #we are allowed upto 2 meals which have been made in past week for 10 meals/week
            if (not inPOrds(meal, menuName)):
                if (inLastWeek(meal, menuName, 10)):
                    numMealsInLastWeek += 1

                if (numMealsInLastWeek > 2):
                    continue
                POrds.at[completedMeals, menuName] = meal
                completedMeals += 1
            
    pickRandomMeal('Asian', 'A', numAsian, completedAsian)
    pickRandomMeal('Classic', 'C', numClassics, completedClassics)

    # ----------------------------------------------------------------------
    # Add the custom order quatities to the POrds quantity list, else record as custom order
    global Customs
    Customs = pd.DataFrame(np.zeros((len(customOrders), 2)), columns=['Customs', 'numCust'], dtype=object)

    global numCustom
    numCustom = 0
    for i, meal in enumerate(customOrders):
        custOrderMenuType = ('numC', 'numA')[mealMenu[meal] == 'A']
        custOrderMenuName = ('Classic', 'Asian')[mealMenu[meal] == 'A']
        added = False

        #Search through list and add quantity
        for j in range(10):    
            if (meal == POrds.at[j, custOrderMenuName] and POrds.at[j, custOrderMenuType] != 0):
                POrds.at[j, custOrderMenuType] += customOrdersQuantity[i]
                added = True
                break
        if (not added):
            for m, k in enumerate(Customs['Customs']):
                if (k == meal):
                    Customs.at[m, 'numCust'] += customOrdersQuantity[i]
                    added = True
            if (not added):    
                Customs.at[numCustom, 'numCust'] = customOrdersQuantity[i]
                Customs.at[numCustom, 'Customs'] = meal
                numCustom += 1
                                
                

    print(POrds)
    print(Customs)

    # ----------------------------------------------------------------------
    # Save POrds to an excel sheet so we can look back on historical data
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                           truncate_sheet=False, header=True,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError
        try:
            # try to open an existing workbook
            writer.book = oxl.load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0
        if startcol is None:
            startcol = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs, header=header)
        # save the workbook
        writer.save()
        
    #Call the function
    append_df_to_excel(getCurrentWeek() + '.xlsx', POrds, startrow=0, index=False)
    append_df_to_excel(getCurrentWeek() + '.xlsx', Customs, startrow=13, index=False)


    # ----------------------------------------------------------------------
    # Write to 'Ingredient and Meal costing1.xlsx'

    IngredientExcel = pd.read_excel(INGREDIENTS_FILE, sheet_name='Chosen Meals')
    

    for i in range(2, 45):
        IngredientExcel.at[i, 'Unnamed: 3'] = 0
    quantities = IngredientExcel['Unnamed: 3']
    meals = IngredientExcel['Unnamed: 2']

    #iterate through POrds adding quantities
    def add_quantities_to_Ingredient_Sheet(mealType, mealQuantity):
        for i in range(10):
            for j in range(2, 45):
                if (POrds.at[i, mealType] == meals.at[j]):
                    quantities.at[j] = POrds.at[i, mealQuantity]

    add_quantities_to_Ingredient_Sheet('Classic', 'numC')
    add_quantities_to_Ingredient_Sheet('Asian', 'numA')

    for a, b in enumerate(Customs['Customs']):
        for j in range(2, 45):
            if (b == meals.at[j]):
                quantities.at[j] = Customs.at[a, 'numCust']

    append_df_to_excel(INGREDIENTS_FILE, quantities, sheet_name='Chosen Meals', startcol=3, startrow=1, index=False, header=False)

if __name__ == '__main__':
    main()
    
